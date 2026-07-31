"""
cuadros_lib.py
================
Lógica de cálculo y generación de los Cuadros de Subejercicio (DGAPSGB).
Usada por app.py (Streamlit). Sin dependencias de Colab / consola.
"""

import re
import datetime
import calendar
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════════════════════
# CATÁLOGOS
# ═══════════════════════════════════════════════════════════════════════════════

MESES_COLS = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEP','OCT','NOV','DIC']

MES_ES       = {1:'enero',2:'febrero',3:'marzo',4:'abril',5:'mayo',6:'junio',
                7:'julio',8:'agosto',9:'septiembre',10:'octubre',11:'noviembre',12:'diciembre'}
MES_ES_ABREV = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
                7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
MES_ES_LARGO = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

UR_MAP = {
    100:100,106:106,107:107,110:110,111:111,112:112,117:117,
    119:119,810:119,812:119,
    120:120,800:120,811:120,
    200:200,220:220,221:221,222:222,
    235:250,250:250,252:252,236:253,253:253,237:253,
    260:260,261:261,262:262,263:263,264:264,265:265,266:266,267:267,
    268:268,269:269,270:270,271:271,272:272,273:273,274:274,275:275,
    276:276,277:277,278:278,279:279,280:280,281:281,282:282,283:283,
    284:284,285:285,286:286,287:287,288:288,289:289,290:290,291:291,292:292,
    500:500,510:510,511:511,512:512,513:513,
    225:900,900:900,245:910,910:910,241:911,911:911,
    246:912,247:912,912:912,
    230:920,920:920,226:921,921:921,227:922,922:922,
    231:923,923:923,924:924,232:924,
    228:'NA - 228',233:'NA - 233',240:'NA - 240',242:'NA - 242',251:'NA - 251',
    'A1I':'A1I','AFU':'AFU','B00':'B00','C00':'C00','D00':'D00','I00':'I00',
    'I6L':'I6L','I9H':'I9H','IZC':'IZC','IZI':'IZI','JAG':'JAG','JAL':'JAL',
    'RJL':'RJL','VSS':'VSS','VST':'VST',
}

CAP_NOMBRES = {
    1000:'Servicios personales',
    2000:'Materiales y suministros',
    3000:'Servicios generales',
    4000:'Transferencias, asignaciones, subsidios y otras ayudas',
    7000:'Inversión pública',
}

PP_NOMBRES = {
    'Q004':'Desarrollo y aplicación de programas y proyectos educativos y de investigación en el sector agroalimentario',
    'P021':'Aplicación de la Política Agropecuaria',
    'M001':'Actividades de apoyo administrativo',
    'O001':'Actividades de apoyo a la función pública y buen gobierno',
    'S263':'Apoyo al ingreso agropecuario: PROCAMPO para vivir mejor',
    'S293':'Producción para el Bienestar',
    'S292':'Fertilizantes para el Bienestar',
    'B006':'Prevención y manejo de riesgos',
    'S052':'Programa de Apoyos a Pequeños Productores',
    'S304':'Pesca y Acuacultura Sustentables',
    'K017':'Infraestructura para el desarrollo rural sustentable',
    'S053':'Programa de Fomento a la Agricultura',
    'S290':'Sembrando Vida',
    'S318':'Comercio Justo',
}

UR_NOMBRES = {
    100:'Secretaría de Agricultura y Desarrollo Rural',110:'Oficialía Mayor',
    111:'Dirección General de Programación, Organización y Presupuesto',
    112:'Dirección General Jurídica',117:'Dirección General de Comunicación Social',
    119:'Dirección General de Planeación y Evaluación',
    120:'Dirección General del Servicio de Información Agroalimentaria y Pesquera',
    200:'Subsecretaría de Agricultura',220:'Unidad de Bienestar para el Campo',
    221:'Dirección General de Fertilizantes para el Bienestar',
    222:'Dirección General de Producción para el Bienestar',
    250:'Dirección General de Fomento a la Agricultura',
    252:'Dirección General de Sanidad Vegetal',
    253:'Dirección General de Inocuidad Agroalimentaria, Acuícola y Pesquera',
    260:'Oficina de Representación en Aguascalientes',
    261:'Oficina de Representación en Baja California',
    262:'Oficina de Representación en Baja California Sur',
    263:'Oficina de Representación en Campeche',264:'Oficina de Representación en Coahuila',
    265:'Oficina de Representación en Colima',266:'Oficina de Representación en Chiapas',
    267:'Oficina de Representación en Chihuahua',268:'Oficina de Representación en Durango',
    269:'Oficina de Representación en Guanajuato',270:'Oficina de Representación en Guanajuato',
    271:'Oficina de Representación en Guerrero',272:'Oficina de Representación en Hidalgo',
    273:'Oficina de Representación en Jalisco',274:'Oficina de Representación en Estado de México',
    275:'Oficina de Representación en Michoacán',276:'Oficina de Representación en Morelos',
    277:'Oficina de Representación en Nayarit',278:'Oficina de Representación en Nuevo León',
    279:'Oficina de Representación en Oaxaca',280:'Oficina de Representación en Puebla',
    281:'Oficina de Representación en Querétaro',282:'Oficina de Representación en Quintana Roo',
    283:'Oficina de Representación en San Luis Potosí',284:'Oficina de Representación en Sinaloa',
    285:'Oficina de Representación en Sonora',286:'Oficina de Representación en Tabasco',
    287:'Oficina de Representación en Tamaulipas',288:'Oficina de Representación en Tlaxcala',
    289:'Oficina de Representación en Veracruz',290:'Oficina de Representación en Yucatán',
    291:'Oficina de Representación en Zacatecas',292:'Oficina de Representación en Ciudad de México',
    500:'Subsecretaría de Pesca y Acuacultura',510:'Dirección General de Acuacultura',
    511:'Dirección General de Ordenamiento Pesquero y Acuícola',
    512:'Dirección General de Recursos Materiales, Inmuebles y Servicios',
    513:'Dirección General de Tecnologías de la Información y Comunicaciones',
    900:'Dirección General de Precios, Ordenamiento Comercial y Valor Agregado',
    910:'Dirección General de Fibras Naturales y Empresas de Participación Estatal',
    911:'Dirección General de Normatividad, Innovación y Sustentabilidad Agropecuaria',
    912:'Dirección General de Productividad y Desarrollo Tecnológico',
    920:'Dirección General de Ganadería',921:'Dirección General de Sanidad Animal',
    922:'Dirección General de Producción Ganadera, Pesquera y Acuícola',
    923:'Dirección General de Precios, Ordenamiento Comercial y Valor Agregado',
    924:'Dirección General de Ordenamiento y Comercialización Agropecuaria',
    'A1I':'Universidad Autónoma Chapingo',
    'AFU':'Comité Nacional para el Desarrollo Sustentable de la Caña de Azúcar',
    'B00':'Servicio Nacional de Sanidad, Inocuidad y Calidad Agroalimentaria',
    'C00':'Servicio Nacional de Inspección y Certificación de Semillas',
    'D00':'Colegio Superior Agropecuario del Estado de Guerrero',
    'I00':'Comisión Nacional de Acuacultura y Pesca',
    'I6L':'Fideicomiso de Riesgo Compartido',
    'I9H':'Instituto Nacional para el Desarrollo de Capacidades del Sector Rural, A.C.',
    'IZC':'Colegio de Postgraduados','IZI':'Comisión Nacional de las Zonas Áridas',
    'JAG':'Instituto Nacional de Investigaciones Forestales, Agrícolas y Pecuarias',
    'JAL':'Junta de Asistencia Privada del Estado de Jalisco',
    'RJL':'Instituto Mexicano de Investigación en Pesca y Acuacultura Sustentables',
    'VSS':'Seguridad Alimentaria Mexicana','VST':'Productora Nacional de Biológicos Veterinarios',
    'NA - 228':'UR 228','NA - 233':'UR 233','NA - 240':'UR 240',
    'NA - 242':'UR 242','NA - 251':'UR 251',
}

# ═══════════════════════════════════════════════════════════════════════════════
# LECTURA Y CÁLCULO
# ═══════════════════════════════════════════════════════════════════════════════

def parse_fecha(fname):
    """Detecta corte desde nombre '01-ABRIL-2026_MAP...' → date(2026,3,31)."""
    meses = {'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,
             'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12,
             'ENERO':1,'FEBRERO':2,'MARZO':3,'ABRIL':4,'MAYO':5,'JUNIO':6,
             'JULIO':7,'AGOSTO':8,'SEPTIEMBRE':9,'OCTUBRE':10,'NOVIEMBRE':11,'DICIEMBRE':12}
    m = re.search(r'(\d{1,2})[_\-]([A-Z]+)[_\-](\d{4})', fname.upper())
    if not m:
        return None
    day, mes_s, year = int(m.group(1)), m.group(2), int(m.group(3))
    mes_num = meses.get(mes_s)
    if not mes_num:
        return None
    fd = datetime.date(year, mes_num, day)
    if fd.day == 1:
        pm = fd.month - 1 if fd.month > 1 else 12
        py = fd.year if fd.month > 1 else fd.year - 1
        return datetime.date(py, pm, calendar.monthrange(py, pm)[1])
    return fd


def _leer_raw(file_obj, filename):
    """file_obj: ruta (str) o objeto tipo-archivo (BytesIO / UploadedFile)."""
    ext = filename.lower().rsplit('.', 1)[-1]
    if ext == 'csv':
        for enc in ['latin-1', 'utf-8', 'utf-8-sig']:
            try:
                if hasattr(file_obj, 'seek'):
                    file_obj.seek(0)
                return pd.read_csv(file_obj, encoding=enc, low_memory=False)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"No se pudo leer {filename}")
    # XLSX: buscar hoja con columnas crudas
    import openpyxl as ox
    if hasattr(file_obj, 'seek'):
        file_obj.seek(0)
    wb = ox.load_workbook(file_obj, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdr = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        if 'UNIDAD' in hdr and 'IDEN_PROY' in hdr and 'PARTIDA' in hdr:
            rows = list(ws.iter_rows(min_row=3, values_only=True))
            wb.close()
            return pd.DataFrame(rows, columns=hdr)
    wb.close()
    raise ValueError("No se encontró hoja MAP en el XLSX (se esperaban columnas UNIDAD / IDEN_PROY / PARTIDA).")


def leer_y_calcular(file_obj, filename, corte_mes):
    df = _leer_raw(file_obj, filename)

    def map_ur(u):
        try:
            return UR_MAP.get(int(u), str(u))
        except (ValueError, TypeError):
            return UR_MAP.get(str(u).strip(), str(u).strip())

    df['UR2'] = df['UNIDAD'].apply(map_ur)
    df['PP'] = df['IDEN_PROY'].astype(str).str.strip() + df['PROYECTO'].astype(str).str.zfill(3)
    df['CAPITULO'] = (df['PARTIDA'].astype(str).str[0] + '000').astype(int)

    sufijos = MESES_COLS[:corte_mes]
    cols = {
        'ORI': [f'ORI_{s}' for s in sufijos if f'ORI_{s}' in df.columns],
        'MOD': [f'MOD_{s}' for s in sufijos if f'MOD_{s}' in df.columns],
        'EJE': [f'EJE_{s}' for s in sufijos if f'EJE_{s}' in df.columns],
    }
    result = pd.DataFrame({
        'UR2':      df['UR2'],
        'PP':       df['PP'],
        'CAPITULO': df['CAPITULO'],
        'ORI':      df[cols['ORI']].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1) / 1e6,
        'MOD':      df[cols['MOD']].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1) / 1e6,
        'EJE':      df[cols['EJE']].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1) / 1e6,
    })
    result['DISP'] = result['MOD'] - result['EJE']
    return result


def agregar(df):
    return df.groupby(['UR2', 'PP', 'CAPITULO'], as_index=False).agg(
        ORI=('ORI', 'sum'), MOD=('MOD', 'sum'), EJE=('EJE', 'sum'), DISP=('DISP', 'sum'))


def seleccionar(grp, disp_min, capitulo_max, lista_manual):
    """
    Un capítulo califica si CAPITULO <= capitulo_max (gasto corriente, excluye
    Inversión >=5000 por default) y su disponible absoluto es >= disp_min.
    Un UR+PP se incluye si le queda al menos un capítulo tras el filtro.
    """
    caps_ok = grp[(grp['CAPITULO'] <= capitulo_max) & (grp['DISP'] >= disp_min)].copy()
    if lista_manual:
        def _norm(u):
            try:
                return UR_MAP.get(int(u), str(u))
            except (ValueError, TypeError):
                return UR_MAP.get(str(u), str(u))
        norm = {(_norm(u), pp) for u, pp in lista_manual}
        mask = caps_ok.apply(lambda r: (r['UR2'], r['PP']) in norm, axis=1)
        return caps_ok[mask].copy()
    return caps_ok


def resumen_general(caps_sel):
    """Tabla UR+PP con MOD/EJE/DISP/PCT, para vista previa antes de descargar."""
    if caps_sel.empty:
        return pd.DataFrame(columns=['UR', 'UR_Nombre', 'PP', 'PP_Nombre', 'MOD_MDP', 'EJE_MDP', 'DISP_MDP', 'PCT'])
    resumen = caps_sel.groupby(['UR2', 'PP'], as_index=False).agg(
        MOD=('MOD', 'sum'), EJE=('EJE', 'sum'), DISP=('DISP', 'sum'))
    resumen['PCT'] = resumen.apply(lambda r: (r['DISP'] / r['MOD'] * 100) if r['MOD'] else 0, axis=1)
    resumen['UR_Nombre'] = resumen['UR2'].apply(lambda u: UR_NOMBRES.get(u, str(u)))
    resumen['PP_Nombre'] = resumen['PP'].apply(lambda p: PP_NOMBRES.get(p, p))
    resumen = resumen.rename(columns={'UR2': 'UR', 'MOD': 'MOD_MDP', 'EJE': 'EJE_MDP', 'DISP': 'DISP_MDP'})
    resumen = resumen[['UR', 'UR_Nombre', 'PP', 'PP_Nombre', 'MOD_MDP', 'EJE_MDP', 'DISP_MDP', 'PCT']]
    resumen = resumen.sort_values('DISP_MDP', ascending=False).reset_index(drop=True)
    return resumen


# ═══════════════════════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════════════════════

NF = '_-\\ #,##0.0_-;[Red]\\-\\ #,##0.0_-;_-\\ "-"??_-;_-@_-'

BURGUNDY = 'FF9D2449'
CREAM = 'FFF6F2EB'


def _f(bold=False, size=10, color='FF000000'):
    return Font(bold=bold, size=size, color=color, name='Arial')


def _fill(rgb):
    return PatternFill('solid', fgColor=rgb) if rgb else PatternFill(fill_type=None)


def _al(h='left', v='top', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _S(style):
    return Side(style=style) if style else Side(style=None)


def _bd(**kw):
    return Border(top=_S(kw.get('t')), bottom=_S(kw.get('b')),
                  left=_S(kw.get('l')), right=_S(kw.get('r')))


def _c(ws, row, col, val=None, bold=False, size=10, color='FF000000',
       h='left', v='top', wrap=False, fill=None, bd=None, nf=None):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    c.font = _f(bold=bold, size=size, color=color)
    c.alignment = _al(h=h, v=v, wrap=wrap)
    if fill is not None:
        c.fill = _fill(fill)
    if bd is not None:
        c.border = bd
    if nf is not None:
        c.number_format = nf
    return c


def _mg(ws, r1, c1, r2, c2):
    try:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE HOJA / WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

def generar_hoja(wb, ur, pp, caps_df, anio, periodo_label, fecha_corte_str, fecha_rep_str):
    ur_key = int(ur) if str(ur).lstrip('-').isdigit() else ur
    ur_nombre = UR_NOMBRES.get(ur_key, str(ur))
    pp_nombre = PP_NOMBRES.get(pp, pp)

    sname = f'Pp {pp} - {ur}'[:31]
    if sname in wb.sheetnames:
        sname = f'{sname[:28]}_{len(wb.sheetnames)}'
    ws = wb.create_sheet(title=sname)

    for ltr, w in [('A', 0.86), ('B', 1.71), ('C', 3.71), ('D', 8.0), ('E', 45.71),
                    ('F', 13.71), ('G', 13.71), ('H', 13.71), ('I', 1.71), ('J', 13.71),
                    ('K', 10.0), ('L', 65.71), ('M', 0.86), ('N', 11.43)]:
        ws.column_dimensions[ltr].width = w

    ws.row_dimensions[1].height = 19.5
    ws.row_dimensions[4].height = 18.0
    ws.row_dimensions[9].height = 18.0

    _c(ws, 1, 2, 'Gasto programable DGAPSG"B"', bold=True, size=11, v='top')
    _c(ws, 2, 2, f'Avance del gasto ejercido por Programa Presupuestario al {fecha_corte_str} (Principales subejercicios)',
       bold=True, size=10, h='left', v='top')
    _c(ws, 3, 2, 'Millones de pesos', size=10, h='left', v='top')

    _c(ws, 5, 2, 'Ramo / PP', bold=True, size=10, h='center', v='center', bd=_bd(t='medium', b='medium'))
    _c(ws, 5, 6, str(anio), bold=True, size=10, h='center', v='center', bd=_bd(t='medium', b='thin'))
    _c(ws, 5, 10, 'Disponible al periodo', bold=True, size=10, h='center', v='center', bd=_bd(t='medium', b='thin'))
    _c(ws, 5, 12, 'Comentarios', bold=True, size=10, h='center', v='center', bd=_bd(t='medium', b='medium'))
    _mg(ws, 5, 2, 9, 5); _mg(ws, 5, 6, 5, 8); _mg(ws, 5, 10, 6, 11); _mg(ws, 5, 12, 9, 12)

    _c(ws, 6, 6, periodo_label, bold=True, size=10, h='center', v='center', bd=_bd(t='thin', b='thin'))
    _mg(ws, 6, 6, 6, 8)

    for col, label in [(6, 'Aprobado'), (7, 'Modificado'), (8, 'Ejercido'), (10, 'Absoluto'), (11, '%')]:
        _c(ws, 7, col, label, bold=True, size=10, h='center', v='center', bd=_bd(t='thin'))
    _mg(ws, 7, 6, 8, 6); _mg(ws, 7, 7, 8, 7); _mg(ws, 7, 8, 8, 8)
    _mg(ws, 7, 10, 8, 10); _mg(ws, 7, 11, 8, 11)

    for col, label in [(6, 'a'), (7, 'b'), (8, 'c'), (10, 'd = b - c'), (11, 'e = d / b')]:
        _c(ws, 9, col, label, size=10, h='center', v='center', bd=_bd(b='medium'))

    for col in range(1, 15):
        ws.cell(row=11, column=col).fill = _fill(BURGUNDY)
        ws.cell(row=11, column=col).font = _f(size=9)
    ur_disp = str(int(ur_key)) if isinstance(ur_key, int) else str(ur_key)
    _c(ws, 11, 4, ur_disp, bold=True, size=9, color='FFFFFFFF', h='center', v='top', fill=BURGUNDY)
    _c(ws, 11, 5, ur_nombre, bold=True, size=9, color='FFFFFFFF', h='left', v='top', fill=BURGUNDY)

    ws.row_dimensions[14].height = 18.0
    for col in range(1, 15):
        ws.cell(row=14, column=col).fill = _fill(CREAM)
        ws.cell(row=14, column=col).font = _f(size=8)
    _c(ws, 14, 3, '08', bold=True, size=8, h='center', v='top', fill=CREAM, bd=_bd(b='medium'))
    _c(ws, 14, 4, 'Agricultura y\xa0Desarrollo Rural', bold=True, size=8, v='top', fill=CREAM, bd=_bd(b='medium'))
    _c(ws, 14, 10, '=+G14-H14', bold=True, size=8, h='right', v='top', fill=CREAM, bd=_bd(b='medium'), nf=NF)
    _c(ws, 14, 11, '=+IF(G14=0,"n.a.",IF(ABS((J14/G14)*100)>500,"-o-",((J14/G14)*100)))',
       bold=True, size=8, h='right', v='top', fill=CREAM, bd=_bd(b='medium'), nf=NF)

    caps_sorted = caps_df.sort_values('CAPITULO').reset_index(drop=True)
    n_caps = len(caps_sorted)
    R_CAP_FIRST = 16
    R_CAP_LAST = 15 + n_caps

    for col in range(1, 15):
        ws.cell(row=15, column=col).fill = _fill(CREAM)
        ws.cell(row=15, column=col).font = _f(size=8)
    _c(ws, 15, 4, pp, bold=True, size=8, v='top', fill=CREAM, bd=_bd(b='medium'))
    _c(ws, 15, 5, pp_nombre, bold=True, size=8, v='top', fill=CREAM, bd=_bd(b='medium'))
    for col, let in [(6, 'F'), (7, 'G'), (8, 'H')]:
        _c(ws, 15, col, f'=SUM({let}{R_CAP_FIRST}:{let}{R_CAP_LAST})',
           bold=True, size=8, h='right', v='top', fill=CREAM, bd=_bd(b='medium'), nf=NF)
    _c(ws, 15, 10, f'=SUM(J{R_CAP_FIRST}:J{R_CAP_LAST})',
       bold=True, size=8, h='right', v='top', fill=CREAM, bd=_bd(b='medium'), nf=NF)
    _c(ws, 15, 11, f'=SUM(K{R_CAP_FIRST}:K{R_CAP_LAST})',
       bold=True, size=8, h='right', v='top', fill=CREAM, bd=_bd(b='medium'), nf=NF)

    bh = _bd(t='hair', b='hair')
    for idx, row_data in caps_sorted.iterrows():
        r = R_CAP_FIRST + idx
        cap = int(row_data['CAPITULO'])
        _c(ws, r, 4, cap, size=9, h='center', v='top', bd=bh)
        _c(ws, r, 5, CAP_NOMBRES.get(cap, str(cap)), size=9, h='left', v='top', wrap=True, bd=bh)
        _c(ws, r, 6, round(float(row_data['ORI']), 6), size=9, h='right', v='top', bd=bh, nf=NF)
        _c(ws, r, 7, round(float(row_data['MOD']), 6), size=9, h='right', v='top', bd=bh, nf=NF)
        _c(ws, r, 8, round(float(row_data['EJE']), 6), size=9, h='right', v='top', bd=bh, nf=NF)
        _c(ws, r, 10, f'=+G{r}-H{r}', size=9, h='right', v='top', bd=bh, nf=NF)
        _c(ws, r, 11, f'=+IF(G{r}=0,"n.a.",IF(ABS((J{r}/G{r})*100)>500,"-o-",((J{r}/G{r})*100)))',
           size=9, h='right', v='top', bd=bh, nf=NF)

    ws.row_dimensions[R_CAP_LAST].height = 18.0
    _c(ws, R_CAP_LAST + 2, 12, fecha_rep_str, size=9, h='right', v='top')

    return sname


def generar_workbook(caps_sel, anio, periodo_label, fecha_corte_str, fecha_rep_str):
    """Genera el workbook completo (una hoja por UR+PP) y regresa bytes (.xlsx)."""
    combos = sorted(
        caps_sel[['UR2', 'PP']].drop_duplicates().values.tolist(),
        key=lambda x: (x[1], str(x[0]))
    )
    wb = Workbook()
    wb.remove(wb.active)
    for ur, pp in combos:
        subset = caps_sel[(caps_sel['UR2'] == ur) & (caps_sel['PP'] == pp)]
        if subset.empty:
            continue
        generar_hoja(wb, ur, pp, subset, anio, periodo_label, fecha_corte_str, fecha_rep_str)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), len(wb.sheetnames)


def periodo_info(corte):
    """A partir de la fecha de corte (último día del mes), regresa año, mes,
    etiqueta de periodo, fecha de corte en texto y fecha de reporte."""
    cm = corte.month
    anio = corte.year
    periodo_label = MES_ES_LARGO[1] if cm == 1 else f'{MES_ES_LARGO[1]} - {MES_ES_LARGO[cm]}'
    fecha_corte_str = f'{corte.day} de {MES_ES[cm]} de {anio}'
    sig = corte + datetime.timedelta(days=1)
    fecha_rep_str = f'{sig.day:02d}-{MES_ES_ABREV[sig.month]}-{sig.year}'
    return cm, anio, periodo_label, fecha_corte_str, fecha_rep_str
